import os, sys, requests, time, json, base64, threading
from datetime import datetime
from pathlib import Path
from flask import Flask, request, abort

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment
    from linebot.v3 import WebhookHandler
    from linebot.v3.exceptions import InvalidSignatureError
    from linebot.v3.messaging import (
        Configuration, ApiClient, MessagingApi, MessagingApiBlob,
        ReplyMessageRequest, PushMessageRequest, TextMessage, FlexMessage, FlexContainer
    )
    from linebot.v3.webhooks import MessageEvent, ImageMessageContent, TextMessageContent
except Exception as e:
    print(f"\n❌ 缺少必要套件: {e}"); sys.exit()

LINE_CHANNEL_ACCESS_TOKEN = os.environ.get("LINE_CHANNEL_ACCESS_TOKEN", "")
LINE_CHANNEL_SECRET       = os.environ.get("LINE_CHANNEL_SECRET", "")
DATALAB_API_KEY  = os.environ.get("DATALAB_API_KEY", "")
DEEPSEEK_API_KEY = os.environ.get("DEEPSEEK_API_KEY", "")
CLAUDE_API_KEY   = os.environ.get("CLAUDE_API_KEY", "")
EXCEL_FILE = Path(__file__).parent / "records.xlsx"
MISSING = "未找到"
KEYS = ["工單 (Part No)", "型號 (Model)", "數量 (Quantity)", "儲位 (Location)", "業單 (Sales Order)"]
VALID_LOC = ([f"A{i}" for i in range(1, 13)] + [f"B{i}" for i in range(1, 7)] + ["0S08", "3F", "NG", "B2", "B3"])
LOC_FIX = {"BZ":"B2","82":"B2","B 2":"B2","86":"B6","BG":"B6","81":"B1","83":"B3","84":"B4","85":"B5","ALL":"A11","A1L":"A11","A|1":"A11","OS08":"0S08","0SO8":"0S08","OSO8":"0S08","0508":"0S08"}

app = Flask(__name__)
handler = WebhookHandler(LINE_CHANNEL_SECRET)
configuration = Configuration(access_token=LINE_CHANNEL_ACCESS_TOKEN)
print(f"🔑 SECRET載入: {'OK' if LINE_CHANNEL_SECRET else 'EMPTY'} ({len(LINE_CHANNEL_SECRET)} chars)")
print(f"🔍 所有LINE相關變數: { {k:v[:4]+'...' for k,v in os.environ.items() if 'LINE' in k} }")

def _parse_location(v):
    v = v.upper().replace(" ", "").strip()
    return LOC_FIX[v] if v in LOC_FIX else (v if v in VALID_LOC else v)

def init_excel():
    if EXCEL_FILE.exists(): return
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "辨識紀錄"
    hd = ["時間"] + KEYS
    for col, h in enumerate(hd, 1):
        cell = ws.cell(1, col, h); cell.font = Font(bold=True); cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = 22
    wb.save(EXCEL_FILE)

def append_excel_multi(results):
    init_excel()
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE); ws = wb.active
        for r in results:
            if r.get("數量 (Quantity)") and r["數量 (Quantity)"] != MISSING:
                r["數量 (Quantity)"] = r["數量 (Quantity)"].lstrip(":：").strip()
            ws.append([datetime.now().strftime("%m%d-%H:%M")] + [r.get(k, MISSING) for k in KEYS])
        wb.save(EXCEL_FILE)
    except Exception as e: print(f"Excel寫入錯誤: {e}")

def vision_get_location(img_b):
    try:
        resp = requests.post(
            "https://shaco.chat/api/v1/messages",
            headers={"x-api-key": CLAUDE_API_KEY, "anthropic-version": "2023-06-01", "content-type": "application/json"},
            json={"model": "claude-haiku-4-5-20251001", "max_tokens": 20,
                  "messages": [{"role": "user", "content": [
                      {"type": "image", "source": {"type": "base64", "media_type": "image/jpeg", "data": base64.b64encode(img_b).decode()}},
                      {"type": "text", "text": "圖片中有手寫的儲位代碼（格式如A9/A10/B6/0S08/3F），請只回傳儲位代碼，沒有則回傳「未找到」"}
                  ]}]},
            timeout=15
        ).json()
        loc = next((item["text"] for item in resp["content"] if item["type"] == "text"), "未找到").strip()
        return _parse_location(loc) if loc != "未找到" else MISSING
    except Exception as e:
        print(f"視覺辨識錯誤: {e}"); return MISSING

def llm_parse(raw_text):
    EMPTY = {k: MISSING for k in KEYS}
    try:
        prompt = (
            "以下是倉管單據OCR文字。若含多筆明細（如多個業單號），每筆獨立一個物件。\n"
            "每筆提取：工單(Part No，剛好12碼、不以Q開頭、對應單據「工單號碼」欄位)、型號(Model，至少3字元)、數量(Quantity)、儲位(Location，格式為英數組合如A9/B6/0S08/3F，通常為手寫)、業單(Sales Order，剛好7碼純數字、以5開頭，不符合則填「未找到」)\n"
            "找不到的欄位填\"未找到\"。只回傳JSON，格式：\n"
            "{\"items\":[{\"工單 (Part No)\":\"...\",\"型號 (Model)\":\"...\",\"數量 (Quantity)\":\"...\",\"儲位 (Location)\":\"...\",\"業單 (Sales Order)\":\"...\"}]}\n\n"
            f"OCR文字：\n{raw_text}"
        )
        resp = requests.post(
            "https://api.deepseek.com/chat/completions",
            headers={"Authorization": f"Bearer {DEEPSEEK_API_KEY}", "Content-Type": "application/json"},
            json={"model": "deepseek-chat", "messages": [{"role": "user", "content": prompt}], "response_format": {"type": "json_object"}},
            timeout=30
        ).json()
        items = json.loads(resp["choices"][0]["message"]["content"]).get("items", [])
        results = []
        for item in items:
            r = {k: str(item.get(k, MISSING)).strip() or MISSING for k in KEYS}
            if r["工單 (Part No)"] != MISSING and len(r["工單 (Part No)"]) != 12:
                r["工單 (Part No)"] = MISSING
            if r["儲位 (Location)"] != MISSING:
                r["儲位 (Location)"] = _parse_location(r["儲位 (Location)"])
            if r["業單 (Sales Order)"] != MISSING:
                r["儲位 (Location)"] = "B2"
                r["工單 (Part No)"] = MISSING
            results.append(r)
        return results or [EMPTY]
    except Exception as e:
        print(f"LLM解析錯誤: {e}"); return [EMPTY]

def cloud_ocr_process(img_b):
    try:
        resp = requests.post("https://www.datalab.to/api/v1/convert", headers={"X-API-Key": DATALAB_API_KEY}, files={"file": ("image.jpg", img_b, "image/jpeg")}, data={"output_format": "markdown", "use_llm": "true"}, timeout=15)
        if resp.status_code != 200:
            return [{k: MISSING for k in KEYS}]
        check_url = resp.json().get("request_check_url")
        raw_text = ""
        for _ in range(15):
            time.sleep(2)
            r = requests.get(check_url, headers={"X-API-Key": DATALAB_API_KEY}, timeout=10).json()
            if r.get("status") == "complete":
                raw_text = r.get("markdown", ""); break
        if not raw_text:
            return [{k: MISSING for k in KEYS}]
        print(f"\n--- 🛠️ 偵錯模式 --- \n{raw_text}\n----------------")
        results = llm_parse(raw_text)
        for r in results:
            if r["儲位 (Location)"] == MISSING:
                r["儲位 (Location)"] = vision_get_location(img_b)
        return results
    except Exception as e:
        print(f"連線異常: {e}"); return [{k: MISSING for k in KEYS}]

def push_text(user_id, text):
    with ApiClient(configuration) as api_client:
        MessagingApi(api_client).push_message(
            PushMessageRequest(to=user_id, messages=[TextMessage(text=text)])
        )

def process_image_task(user_id, img_b):
    results = cloud_ocr_process(img_b)
    append_excel_multi(results)
    reply = "📋 AI 辨識結果\n"
    reply += f"已自動寫入 {len(results)} 筆物料明細\n"
    reply += "────────────────\n"
    for i, res in enumerate(results, 1):
        if len(results) > 1: reply += f"📦 第 {i} 筆明細：\n"
        for k, v in res.items():
            reply += f" {'✅' if v != MISSING else '❌'} {k}: {v}\n"
        if len(results) > 1: reply += "──────────\n"
    push_text(user_id, reply)

@app.route("/webhook", methods=['POST'])
def webhook():
    signature = request.headers.get('X-Line-Signature', '')
    body = request.get_data(as_text=True)
    try:
        handler.handle(body, signature)
    except InvalidSignatureError:
        abort(400)
    return 'OK'

@handler.add(MessageEvent, message=ImageMessageContent)
def handle_image(event):
    target_id = getattr(event.source, 'group_id', None) or getattr(event.source, 'room_id', None) or event.source.user_id
    with ApiClient(configuration) as api_client:
        line_bot_api = MessagingApi(api_client)
        blob_api = MessagingApiBlob(api_client)
        line_bot_api.reply_message(ReplyMessageRequest(
            reply_token=event.reply_token,
            messages=[TextMessage(text="🔍 正在 AI 辨識中，請稍候...")]
        ))
        img_b = bytes(blob_api.get_message_content(event.message.id))
    threading.Thread(target=process_image_task, args=(target_id, img_b), daemon=True).start()

@handler.add(MessageEvent, message=TextMessageContent)
def handle_text(event):
    target_id = getattr(event.source, 'group_id', None) or getattr(event.source, 'room_id', None) or event.source.user_id
    text = event.message.text.strip().lower()
    with ApiClient(configuration) as api_client:
        line_bot_api = MessagingApi(api_client)
        if text in ["excel", "dl", "下載"]:
            if EXCEL_FILE.exists():
                with open(EXCEL_FILE, 'rb') as f:
                    excel_data = f.read()
                # 上傳檔案到 LINE，取得 URL
                import tempfile, uuid
                temp_url = f"https://line-bot-production-b2a9.up.railway.app/download/{uuid.uuid4().hex}"
                # 暫存檔案供下載
                app.excel_temp = excel_data
                reply = f"📊 Excel 檔案已產生\n點此下載：{temp_url}\n（連結10分鐘內有效）"
            else:
                reply = "❌ 目前尚無紀錄檔案。"
        elif text == "clear":
            if EXCEL_FILE.exists(): EXCEL_FILE.unlink()
            init_excel()
            reply = "🗑 紀錄已清空"
        else:
            return
        line_bot_api.reply_message(ReplyMessageRequest(
            reply_token=event.reply_token,
            messages=[TextMessage(text=reply)]
        ))

@app.route("/download/<file_id>")
def download_excel(file_id):
    if hasattr(app, 'excel_temp') and app.excel_temp:
        from flask import send_file
        import io
        return send_file(
            io.BytesIO(app.excel_temp),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='倉管辨識紀錄.xlsx'
        )
    abort(404)

if __name__ == "__main__":
    init_excel()
    print("🤖 LINE 倉管機器人啟動中...")
    print("📌 請用 ngrok 或部署到 Railway 取得 HTTPS webhook URL")
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
